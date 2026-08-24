import csv
import io
import re
from pathlib import Path

from django.db import transaction

from turmas.models import Turma

from .models import Aluno


MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_ROWS = 5000
HEADERS = {
    'numero': {'numero', 'número', 'numero do aluno', 'número do aluno', 'nº', 'n.o'},
    'nome': {'nome', 'nome completo'},
    'encarregado_educacao': {'encarregado', 'nome do encarregado', 'nome do encarregado (opcional)', 'encarregado de educacao', 'encarregado de educação'},
    'telefone_encarregado': {'contacto', 'contacto do encarregado', 'contacto do encarregado (opcional)', 'telefone', 'telefone do encarregado'},
}


class SpreadsheetError(ValueError):
    pass


def _normalise_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _read_rows(upload):
    extension = Path(upload.name).suffix.lower()
    if extension not in {'.csv', '.xlsx'}:
        raise SpreadsheetError('Extensão não suportada. Utilize um ficheiro CSV ou XLSX.')
    if upload.size > MAX_FILE_SIZE:
        raise SpreadsheetError('O ficheiro excede o limite de 5 MB.')

    if extension == '.csv':
        try:
            text = upload.read().decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise SpreadsheetError('Não foi possível interpretar o CSV. Utilize codificação UTF-8.') from exc
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
    else:
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            workbook.close()
        except Exception as exc:
            raise SpreadsheetError('O ficheiro XLSX é inválido ou não pode ser interpretado.') from exc

    if not rows:
        raise SpreadsheetError('A planilha está vazia.')
    if len(rows) - 1 > MAX_ROWS:
        raise SpreadsheetError(f'A planilha excede o limite de {MAX_ROWS} linhas.')
    return rows


def analyse_upload(upload, turma_id, mode='importar'):
    try:
        turma = Turma.objects.get(pk=turma_id)
    except (Turma.DoesNotExist, TypeError, ValueError) as exc:
        raise SpreadsheetError('A turma seleccionada não existe.') from exc

    rows = _read_rows(upload)
    header_map = {}
    for index, raw_header in enumerate(rows[0]):
        header = _normalise_header(raw_header)
        for field, aliases in HEADERS.items():
            if header in aliases:
                header_map[field] = index
                break
    if 'numero' not in header_map or 'nome' not in header_map:
        raise SpreadsheetError('A planilha deve conter as colunas Número do aluno e Nome completo.')

    parsed = []
    counts = {}
    for line_number, raw_row in enumerate(rows[1:], start=2):
        values = [str(value).strip() if value is not None else '' for value in raw_row]
        if not any(values):
            parsed.append({'linha': line_number, 'numero': '', 'nome': '', 'estado': 'Erro', 'erros': ['Linha vazia.']})
            continue

        def cell(field):
            index = header_map.get(field)
            return values[index] if index is not None and index < len(values) else ''

        raw_number = cell('numero')
        if raw_number.endswith('.0'):
            raw_number = raw_number[:-2]
        errors = []
        number = None
        if not raw_number:
            errors.append('Número do aluno vazio.')
        elif not raw_number.isdigit() or int(raw_number) <= 0:
            errors.append('Número do aluno inválido.')
        else:
            number = int(raw_number)
            counts[number] = counts.get(number, 0) + 1
        nome = cell('nome').strip()
        if not nome:
            errors.append('Nome completo vazio.')

        item = {'linha': line_number, 'numero': number or raw_number, 'nome': nome, 'erros': errors, 'dados': {}}
        for field in ('encarregado_educacao', 'telefone_encarregado'):
            if field in header_map and cell(field):
                item['dados'][field] = cell(field).strip()
        parsed.append(item)

    duplicate_numbers = {number for number, count in counts.items() if count > 1}
    for item in parsed:
        if isinstance(item.get('numero'), int) and item['numero'] in duplicate_numbers:
            item['erros'].append('Número de aluno duplicado na planilha.')

    valid_numbers = [item['numero'] for item in parsed if isinstance(item.get('numero'), int)]
    existing_by_number = {}
    for aluno in Aluno.objects.filter(numero__in=valid_numbers).select_related('turma'):
        existing_by_number.setdefault(aluno.numero, []).append(aluno)

    summary = {'novos': 0, 'actualizacoes': 0, 'sem_alteracoes': 0, 'erros': 0}
    for item in parsed:
        if item['erros']:
            item['estado'] = 'Erro'
            summary['erros'] += 1
            continue
        matches = existing_by_number.get(item['numero'], [])
        if len(matches) > 1:
            item['estado'] = 'Erro'
            item['erros'].append('Existem vários alunos com este número no SIGEP; corrija os dados antes de continuar.')
            summary['erros'] += 1
        elif not matches:
            if mode == 'actualizar':
                item['estado'] = 'Erro'
                item['erros'].append('Aluno não encontrado para actualização.')
                summary['erros'] += 1
            else:
                item['estado'] = 'Novo'
                summary['novos'] += 1
        else:
            aluno = matches[0]
            changes = {}
            candidate = {'nome': item['nome'], 'turma_id': turma.id, **item['dados']}
            for field, new_value in candidate.items():
                old_value = getattr(aluno, field)
                if old_value != new_value:
                    changes[field] = {'actual': old_value, 'novo': new_value}
            item['aluno_id'] = aluno.id
            item['alteracoes'] = changes
            item['estado'] = 'Actualização' if changes else 'Sem alterações'
            summary['actualizacoes' if changes else 'sem_alteracoes'] += 1
    return {'turma': {'id': turma.id, 'nome': str(turma)}, 'resumo': summary, 'linhas': parsed, '_turma': turma}


@transaction.atomic
def confirm_upload(upload, turma_id, mode='importar'):
    analysis = analyse_upload(upload, turma_id, mode)
    if analysis['resumo']['erros']:
        raise SpreadsheetError('Corrija os erros da planilha antes de confirmar a importação.')
    turma = analysis.pop('_turma')
    for item in analysis['linhas']:
        defaults = {'nome': item['nome'], 'turma': turma, **item['dados']}
        if item['estado'] == 'Novo':
            Aluno.objects.create(numero=item['numero'], **defaults)
        elif item['estado'] == 'Actualização':
            aluno = Aluno.objects.select_for_update().get(pk=item['aluno_id'])
            for field, value in defaults.items():
                setattr(aluno, field, value)
            aluno.save()
    return analysis
